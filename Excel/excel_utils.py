import openpyxl


def get_max_row(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["sheetName"]
    return sheet.max_row


def get_max_col(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["sheetName"]
    return sheet.max_column


def get_cell_data(file, sheetName, rowNo, colNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["sheetName"]
    return sheet.cell(rowNo, colNo).value


def set_cell_data(file, sheetName, rowNo, colNo, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["sheetName"]
    sheet.cell(rowNo, colNo).value = data
    workbook.save(file)
